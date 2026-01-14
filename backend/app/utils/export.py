"""
데이터 내보내기 유틸리티

CSV 및 Excel 형식으로 데이터를 내보내는 기능을 제공합니다.
"""

import csv
import io
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_csv(data: List[Dict[str, Any]], columns: List[str] = None) -> str:
    """
    딕셔너리 리스트를 CSV 문자열로 변환

    Args:
        data: 데이터 리스트 (각 항목은 딕셔너리)
        columns: 포함할 컬럼 목록 (None이면 모든 키 사용)

    Returns:
        CSV 형식의 문자열
    """
    if not data:
        return ""

    output = io.StringIO()

    # 컬럼 결정
    if columns is None:
        columns = list(data[0].keys())

    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue()


def generate_excel(
    data: List[Dict[str, Any]],
    columns: List[str] = None,
    sheet_name: str = "Sheet1",
    title: str = None
) -> bytes:
    """
    딕셔너리 리스트를 Excel 바이트로 변환 (스타일 포함)

    Args:
        data: 데이터 리스트
        columns: 포함할 컬럼 목록
        sheet_name: 시트 이름
        title: 상단에 표시할 제목 (선택사항)

    Returns:
        Excel 파일의 바이트 데이터
    """
    if not data:
        # 빈 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 컬럼 결정
    if columns is None:
        columns = list(data[0].keys())

    # DataFrame 생성
    df = pd.DataFrame(data)[columns]

    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_offset = 1

    # 제목 추가 (선택사항)
    if title:
        ws.merge_cells(f'A1:{chr(64 + len(columns))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row_offset = 3  # 제목 + 빈 줄

    # 헤더 작성
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=row_offset, column=col_idx)
        cell.value = column
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 데이터 작성
    for row_idx, row_data in enumerate(data, start=row_offset + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(column, "")

            # datetime 객체를 문자열로 변환
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")

            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 컬럼 너비 자동 조정
    for col_idx, column in enumerate(columns, start=1):
        max_length = len(str(column))
        for row_data in data:
            value = str(row_data.get(column, ""))
            max_length = max(max_length, len(value))

        # 최대 50자로 제한
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[chr(64 + col_idx)].width = adjusted_width

    # 바이트로 변환
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def generate_diagnosis_csv(diagnosis_data: Dict[str, Any]) -> str:
    """
    진단 결과를 CSV로 변환

    Args:
        diagnosis_data: 진단 결과 데이터

    Returns:
        CSV 형식의 문자열
    """
    # 기본 정보
    basic_info = [
        {"항목": "진단 ID", "값": diagnosis_data.get("diagnosis_id", "")},
        {"항목": "투자 성향", "값": diagnosis_data.get("investment_type", "")},
        {"항목": "점수", "값": diagnosis_data.get("score", "")},
        {"항목": "신뢰도", "값": f"{diagnosis_data.get('confidence', 0) * 100:.1f}%"},
        {"항목": "월 투자액", "값": f"{diagnosis_data.get('monthly_investment', 0):,}만원" if diagnosis_data.get('monthly_investment') else "미입력"},
        {"항목": "진단일시", "값": diagnosis_data.get("created_at", "")},
    ]

    output = io.StringIO()

    # 기본 정보
    output.write("=== 진단 기본 정보 ===\n")
    writer = csv.DictWriter(output, fieldnames=["항목", "값"])
    writer.writeheader()
    writer.writerows(basic_info)

    # 특징
    output.write("\n=== 투자 성향 특징 ===\n")
    characteristics = diagnosis_data.get("characteristics", [])
    for i, char in enumerate(characteristics, 1):
        output.write(f"{i}. {char}\n")

    # 시나리오 비율
    output.write("\n=== 시나리오 자산 배분 ===\n")
    scenario_ratio = diagnosis_data.get("scenario_ratio", {})
    ratio_data = [{"자산": k, "비율": f"{v}%"} for k, v in scenario_ratio.items()]
    writer = csv.DictWriter(output, fieldnames=["자산", "비율"])
    writer.writeheader()
    writer.writerows(ratio_data)

    # 기대 수익률
    output.write(f"\n기대 연 수익률: {diagnosis_data.get('expected_annual_return', '')}\n")

    return output.getvalue()


def generate_diagnosis_excel(diagnosis_data: Dict[str, Any]) -> bytes:
    """
    진단 결과를 Excel로 변환 (여러 시트)

    Args:
        diagnosis_data: 진단 결과 데이터

    Returns:
        Excel 파일의 바이트 데이터
    """
    wb = Workbook()

    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. 기본 정보 시트
    ws1 = wb.active
    ws1.title = "기본 정보"

    # 제목
    ws1.merge_cells('A1:B1')
    ws1['A1'] = "투자 성향 진단 결과"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 기본 정보
    basic_info = [
        ["진단 ID", diagnosis_data.get("diagnosis_id", "")],
        ["투자 성향", diagnosis_data.get("investment_type", "")],
        ["점수", diagnosis_data.get("score", "")],
        ["신뢰도", f"{diagnosis_data.get('confidence', 0) * 100:.1f}%"],
        ["월 투자액", f"{diagnosis_data.get('monthly_investment', 0):,}만원" if diagnosis_data.get('monthly_investment') else "미입력"],
        ["진단일시", str(diagnosis_data.get("created_at", ""))],
        ["설명", diagnosis_data.get("description", "")],
    ]

    for row_idx, (label, value) in enumerate(basic_info, start=3):
        ws1.cell(row=row_idx, column=1).value = label
        ws1.cell(row=row_idx, column=1).font = Font(bold=True)
        ws1.cell(row=row_idx, column=1).border = border

        ws1.cell(row=row_idx, column=2).value = value
        ws1.cell(row=row_idx, column=2).border = border

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 50

    # 2. 특징 시트
    ws2 = wb.create_sheet("투자 성향 특징")
    ws2['A1'] = "투자 성향 특징"
    ws2['A1'].font = title_font

    characteristics = diagnosis_data.get("characteristics", [])
    for i, char in enumerate(characteristics, start=3):
        ws2.cell(row=i, column=1).value = char
        ws2.cell(row=i, column=1).border = border

    ws2.column_dimensions['A'].width = 80

    # 3. 자산 배분 시트
    ws3 = wb.create_sheet("학습 시나리오 자산 배분")
    ws3['A1'] = "학습 시나리오 자산 배분"
    ws3['A1'].font = title_font

    ws3.cell(row=3, column=1).value = "자산"
    ws3.cell(row=3, column=1).fill = header_fill
    ws3.cell(row=3, column=1).font = header_font
    ws3.cell(row=3, column=1).border = border

    ws3.cell(row=3, column=2).value = "비율"
    ws3.cell(row=3, column=2).fill = header_fill
    ws3.cell(row=3, column=2).font = header_font
    ws3.cell(row=3, column=2).border = border

    scenario_ratio = diagnosis_data.get("scenario_ratio", {})
    for row_idx, (asset, ratio) in enumerate(scenario_ratio.items(), start=4):
        ws3.cell(row=row_idx, column=1).value = asset
        ws3.cell(row=row_idx, column=1).border = border

        ws3.cell(row=row_idx, column=2).value = f"{ratio}%"
        ws3.cell(row=row_idx, column=2).border = border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15

    # 기대 수익률
    last_row = len(scenario_ratio) + 5
    ws3.cell(row=last_row, column=1).value = "기대 연 수익률"
    ws3.cell(row=last_row, column=1).font = Font(bold=True)
    ws3.cell(row=last_row, column=2).value = diagnosis_data.get("expected_annual_return", "")

    # 바이트로 변환
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
